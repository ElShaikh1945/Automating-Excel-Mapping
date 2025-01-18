import os
import logging
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox

# Set the expiration date
expiration_date = datetime(2024, 8, 5)

# Check if the script has expired
if datetime.now() > expiration_date:
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    messagebox.showerror("Script Expired", "This script has expired and is no longer usable. Please email me at Muhammad.AlShaikh@Outlook.com for further assistance.")
    exit()

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Create a Tkinter root window (this will be hidden)
root = tk.Tk()
root.withdraw()

# Open a file dialog for the user to select the file
filename = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx *.xls")])

if not filename:
    print("No file selected. Exiting.")
    exit()

try:
    wb = load_workbook(filename=filename)
    sheet1 = wb['Sheet1']
    sheet2 = wb['Sheet2']
    sheet3 = wb['Sheet3']
except Exception as e:
    print(f"Error loading workbook: {e}")

# Define column mappings
column_mappings = {
    'Employee Code': (sheet1, 2, sheet3, 4),
    'National ID': (sheet1, 2, sheet3, 5),
    'DOB': (None,'=IF(LEFT(D2,1)="3",MID(D2,2,2)+2000,MID(D2,2,2)+1900)&"-"&MID(D2,4,2)&"-"&MID(D2,6,2)', sheet3, 7),
    'DOJ': (sheet1, 4, sheet3, 13),
    'Effective From Date': (sheet1, 3, sheet3, 10),
    'Service Description': (sheet1, 7, sheet3, 21),
    'Employee Type': (None, 'Contract', sheet3, 6),
    'Supplier Code': (None, '1600009494', sheet3, 18),
    'Supplier Name': (None, 'Ergo', sheet3, 19),
    'Market Unit': (None, 'EGYPT Foods', sheet3, 51),
    'Function': (sheet1, 9, sheet3, 52),
    'Sub Function': (sheet1, 10, sheet3, 53),
    'Business Category': (sheet1, 11, sheet3, 56),
    'Product Line': (sheet1, 12, sheet3, 57),
    'Employee Arabic Name': (sheet1, 1, sheet3, 59),
}

department = [cell.value for col in sheet1.columns for cell in col if col[0].column == 8]
for i, value in enumerate(department):
    sheet3.cell(row=i+1, column=29, value=value)
    sheet3.cell(row=i+1, column=54, value=value)


work_location = [cell.value for col in sheet1.columns for cell in col if col[0].column == 6]
for i, value in enumerate(work_location):
    sheet3.cell(row=i+1, column=22, value=value)
    sheet3.cell(row=i+1, column=55, value=value)

# Copy data from sheet1 to sheet3
for column_name, (source_sheet, source_column, target_sheet, target_column) in column_mappings.items():
    if source_sheet is None:
        for row in range(2, target_sheet.max_row + 1):
            target_sheet.cell(row=row, column=target_column).value = source_column
    else:
        for row in range(2, source_sheet.max_row + 1):
            target_sheet.cell(row=row, column=target_column).value = source_sheet.cell(row=row, column=source_column).value

# Map service codes and descriptions
service_mappings = {
    sheet2.cell(row=row, column=8).value: sheet2.cell(row=row, column=7).value
    for row in range(2, sheet2.max_row + 1)
    if sheet2.cell(row=row, column=7).value and sheet2.cell(row=row, column=8).value
}
max_row_ws2 = sheet2.max_row
max_row_ws3 = sheet3.max_row
for row_ws3 in range(1, max_row_ws3 + 1):
    value_u_ws3 = sheet3.cell(row=row_ws3, column=21).value
    for row_ws2 in range(1, max_row_ws2 + 1):
        value_h_ws2 = sheet2.cell(row=row_ws2, column=8).value
        if value_u_ws3 == value_h_ws2:
            value_g_ws2 = sheet2.cell(row=row_ws2, column=7).value
            sheet3.cell(row=row_ws3, column=20).value = value_g_ws2

# Map service codes and descriptions
service_mappings = {
    sheet2.cell(row=row, column=8).value: sheet2.cell(row=row, column=7).value
    for row in range(2, sheet2.max_row + 1)
    if sheet2.cell(row=row, column=7).value and sheet2.cell(row=row, column=8).value
}

for row_ws3 in range(2, sheet3.max_row + 1):
    service_desc_ws3 = sheet3.cell(row=row_ws3, column=21).value
    for row_ws2 in range(2, sheet2.max_row + 1):
        service_desc_ws2 = sheet2.cell(row=row_ws2, column=8).value
        if service_desc_ws3 == service_desc_ws2:
            sheet3.cell(row=row_ws3, column=16).value = sheet2.cell(row=row_ws2, column=3).value  
            sheet3.cell(row=row_ws3, column=17).value = sheet2.cell(row=row_ws2, column=4).value  
            sheet3.cell(row=row_ws3, column=18).value = sheet2.cell(row=row_ws2, column=19).value  

# Save the workbook
output_filename = filedialog.asksaveasfilename(title="Save Output File", defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
if output_filename:
    wb.save(output_filename)
    messagebox.showinfo("Success", "Print")
    print(f"Excel file has been updated successfully.")
else:
    print("No output file selected. Exiting.")
